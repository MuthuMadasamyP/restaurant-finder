import io
from datetime import datetime

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_BG = "1E3A5F"
HEADER_FG = "FFFFFF"
ROW_ALT = "F0F4F8"
RATING_HIGH = "D4EDDA"
RATING_MEDIUM = "FFF3CD"
RATING_LOW = "F8D7DA"


def export_to_excel(restaurants: list, location: str) -> bytes:
    """Convert restaurant records to a formatted Excel workbook."""
    rows = [
        {
            "Restaurant Name": _val(restaurant, "name"),
            "Rating": _val(restaurant, "rating"),
            "Category": _val(restaurant, "category"),
            "Address": _val(restaurant, "address"),
            "Phone": _val(restaurant, "phone"),
            "Website URL": _val(restaurant, "website"),
        }
        for restaurant in restaurants
    ]
    df = pd.DataFrame(rows)
    df.index = range(1, len(df) + 1)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, index_label="Rank", sheet_name="Restaurants")
        workbook = writer.book
        worksheet = writer.sheets["Restaurants"]

        _style_header(worksheet)
        _style_rows(worksheet, df)
        _auto_size_columns(worksheet)
        _add_info_sheet(workbook, location, len(restaurants))

    return buffer.getvalue()


def _style_header(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor=HEADER_BG)
    header_font = Font(color=HEADER_FG, bold=True, size=11, name="Calibri")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    border = Border(bottom=Side(style="thin", color="CCCCCC"))

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border

    worksheet.row_dimensions[1].height = 22


def _style_rows(worksheet, df: pd.DataFrame) -> None:
    alt_fill = PatternFill("solid", fgColor=ROW_ALT)
    high_fill = PatternFill("solid", fgColor=RATING_HIGH)
    med_fill = PatternFill("solid", fgColor=RATING_MEDIUM)
    low_fill = PatternFill("solid", fgColor=RATING_LOW)
    border = Border(bottom=Side(style="thin", color="E2E8F0"))

    try:
        rating_col_idx = list(df.columns).index("Rating") + 2
    except ValueError:
        rating_col_idx = None

    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=1):
        is_alt = row_idx % 2 == 0
        for cell in row:
            if is_alt:
                cell.fill = alt_fill
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.border = border

        if rating_col_idx:
            rating_cell = worksheet.cell(row=row_idx + 1, column=rating_col_idx)
            rating_val = _parse_float(str(rating_cell.value))
            if rating_val is not None:
                if rating_val >= 4.0:
                    rating_cell.fill = high_fill
                elif rating_val >= 3.0:
                    rating_cell.fill = med_fill
                else:
                    rating_cell.fill = low_fill


def _auto_size_columns(worksheet) -> None:
    for column in worksheet.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        letter = get_column_letter(column[0].column)
        worksheet.column_dimensions[letter].width = min(max_len + 4, 50)


def _add_info_sheet(workbook, location: str, count: int) -> None:
    worksheet = workbook.create_sheet("Info")
    worksheet.column_dimensions["A"].width = 25
    worksheet.column_dimensions["B"].width = 40

    worksheet["A1"] = "Restaurant Finder Export"
    worksheet["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    worksheet.merge_cells("A1:B1")

    rows = [
        ("Location", location),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Restaurants Found", count),
        ("Data Source", "Google Maps"),
    ]
    for row_idx, (label, value) in enumerate(rows, start=3):
        worksheet.cell(row=row_idx, column=1, value=label).font = Font(bold=True, size=11)
        worksheet.cell(row=row_idx, column=2, value=value).font = Font(size=11)


def _val(obj, attr: str) -> str:
    if isinstance(obj, dict):
        return str(obj.get(attr, "N/A"))
    return str(getattr(obj, attr, "N/A"))


def _parse_float(value: str) -> float | None:
    try:
        return float(value.replace(",", "."))
    except (ValueError, AttributeError):
        return None
